import os, sys
import openpyxl, json, re

excel_path = os.environ.get("MSSI_EXCEL_PATH") or (sys.argv[1] if len(sys.argv) > 1 else "")
if not excel_path:
    raise SystemExit("MSSI_EXCEL_PATH is required")
wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb['결과계산(2)']

with open('/workspace/MSSI/questions.js') as f:
    js_content = f.read()

def parse_questions(js):
    qid_text = []

    sections_match = re.search(r'export const SURVEY_SECTIONS\s*=\s*\[(.*?)\];', js, re.DOTALL)
    if not sections_match:
        sections_match = re.search(r'const SURVEY_SECTIONS\s*=\s*\[(.*?)\];', js, re.DOTALL)

    if not sections_match:
        print("ERROR: Could not find SURVEY_SECTIONS")
        return []

    sections_text = sections_match.group(1)

    section_blocks = re.findall(r'\{([^{}]*(?:\{[^{}]*\}[^{}]*)*)\}', sections_text, re.DOTALL)

    for sblock in section_blocks:

        questions_match = re.search(r'questions\s*:\s*\[(.*?)\]', sblock, re.DOTALL)
        if not questions_match:
            continue

        questions_text = questions_match.group(1)

        section_type = re.search(r'type\s*:\s*"([^"]+)"', sblock)
        stype = section_type.group(1) if section_type else "scale"

        if stype in ('pms_special', 'custom_spaq', 'custom_csm', 'custom_audit', 'custom_mdq', 'custom_mssi', 'custom_ersq'):

            qids_in_block = re.findall(r'\{[^}]*id:\s*"([^"]+)"[^}]*text:\s*"([^"]*)"', questions_text)
            for qid, text in qids_in_block:

                section_ids = {'zung_sds','bai','temps_a','mssi','diag_suicide','diag_panic',
                              'diag_agora','diag_social','diag_ocd','diag_gad','miq_t',
                              'k_mdq','bapq','ctq_sf','ipsm','cd_risc','ersq','bis_bas',
                              'audit_k','csm','spaq','asrs','pai_bor','wurs','pms'}
                if qid not in section_ids:
                    qid_text.append((qid, text))
            continue

        qid_in_block = re.findall(r'\{[^}]*id:\s*"([^"]+)"[^}]*text:\s*"([^"]*)"', questions_text)
        for qid, text in qid_in_block:
            qid_text.append((qid, text))

    return qid_text

qid_text_pairs = parse_questions(js_content)
print(f"Extracted {len(qid_text_pairs)} question ID-text pairs from questions.js")

input_col_info = []
for col in range(3, ws.max_column + 1):
    formula = ws.cell(row=4, column=col).value
    if formula and isinstance(formula, str) and formula.startswith('='):

        if 'VLOOKUP' in formula:

            q_text = ws.cell(row=3, column=col).value
            item_num = ws.cell(row=2, column=col).value
            col_letter = openpyxl.utils.get_column_letter(col)
            input_col_info.append({
                'col': col,
                'col_letter': col_letter,
                'item_num': item_num,
                'q_text': str(q_text).strip() if q_text else ''
            })

print(f"Found {len(input_col_info)} input columns (VLOOKUP formula cells)")

def clean_text(text):
    text = re.sub(r'^\d+[\.\t\s]*', '', text).strip()
    text = re.sub(r'\s+', ' ', text)
    return text

qid_to_col = {}
mapped_count = 0
col_idx = 0

for qid, qtext in qid_text_pairs:
    if col_idx >= len(input_col_info):
        print(f"WARNING: More questions than input columns, stopping at {qid}")
        break

    col_info = input_col_info[col_idx]
    col_letter = col_info['col_letter']
    col_num = col_info['col']

    ref_text_clean = clean_text(col_info['q_text'])
    qtext_clean = clean_text(qtext)

    score = 0
    words = qtext_clean.split()
    for w in words:
        if len(w) >= 2 and w in ref_text_clean:
            score += 1

    match_ratio = score / max(len(words), 1)

    if match_ratio < 0.3:
        print(f"  LOW MATCH: {qid} vs col {col_letter}")
        print(f"    Expected: {qtext_clean[:60]}")
        print(f"    Found:    {ref_text_clean[:60]}")

        print(f"    → Mapped positionally (score={match_ratio:.2f})")

    qid_to_col[qid] = col_num
    col_idx += 1
    mapped_count += 1

print(f"\nMapped {mapped_count} questions to columns")

sections = {}
current_section = None
for col in range(3, ws.max_column + 1):
    hdr = ws.cell(row=1, column=col).value
    if hdr and str(hdr).strip():
        current_section = str(hdr).strip()
    if current_section:
        if current_section not in sections:
            sections[current_section] = {'start': col, 'end': col}
        else:
            sections[current_section]['end'] = col

scoring_cols = {}
for col in range(3, ws.max_column + 1):
    formula = ws.cell(row=4, column=col).value
    if formula and isinstance(formula, str) and formula.startswith('=') and 'VLOOKUP' not in formula:
        section_name = None
        for sn, r in sections.items():
            if r['start'] <= col <= r['end']:
                section_name = sn
                break
        if section_name:
            if section_name not in scoring_cols:
                scoring_cols[section_name] = []
            scoring_cols[section_name].append(openpyxl.utils.get_column_letter(col))

output = []
output.append("const QID_TO_COL = {")
for qid, col_num in qid_to_col.items():
    col_letter = openpyxl.utils.get_column_letter(col_num)
    output.append(f"  '{qid}': {col_num},")
output.append("};")

print("\n\n=== Generated Mapping ===")
print('\n'.join(output))

with open('/workspace/MSSI/supabase/qid_col_map.json', 'w') as f:
    json.dump({
        'qid_to_col': qid_to_col,
        'input_columns': len(input_col_info),
        'mapped': mapped_count,
        'sections': {k: {'start': v['start'], 'end': v['end'],
                         'start_letter': openpyxl.utils.get_column_letter(v['start']),
                         'end_letter': openpyxl.utils.get_column_letter(v['end'])}
                     for k, v in sections.items()}
    }, f, ensure_ascii=False, indent=2)

print(f"\nMapping saved to /workspace/MSSI/supabase/qid_col_map.json")
